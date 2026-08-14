"""Bulk-add mailboxes from Exzelon_Emails.xlsx to production via API."""
import sys
import time
import json
import requests
import openpyxl
from collections import Counter

# --- Config ---
API_BASE = "https://ra.partnerwithus.tech/api/v1"
LOGIN_USER = "ali.aitechs@gmail.com"
LOGIN_PASS = "SA@Admin#123"
EXCEL_PATH = r"C:\Users\Anas\Downloads\Exzelon_Emails.xlsx"

# SMTP/IMAP settings per domain (from cPanel screenshot)
# All domains hosted on same cPanel, each uses mail.<domain>
DEFAULT_SMTP_PORT = 465
DEFAULT_IMAP_PORT = 993
DEFAULT_PROVIDER = "smtp"
DEFAULT_AUTH_METHOD = "password"
DAILY_SEND_LIMIT = 30


def get_mail_server(email: str) -> str:
    """Return mail.<domain> for the email's domain."""
    domain = email.split("@")[1].strip().lower()
    return f"mail.{domain}"


def login() -> str:
    """Authenticate and return access token."""
    resp = requests.post(f"{API_BASE}/auth/login", data={
        "username": LOGIN_USER,
        "password": LOGIN_PASS,
    }, timeout=15)
    resp.raise_for_status()
    return resp.json()["access_token"]


def get_existing_mailboxes(headers: dict) -> set:
    """Return set of existing mailbox emails (lowercased)."""
    resp = requests.get(f"{API_BASE}/mailboxes", headers=headers,
                        params={"limit": 1000}, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    items = data.get("items", []) if isinstance(data, dict) else data
    return {m["email"].lower() for m in items}


def read_excel() -> list:
    """Read all mailbox rows from the Excel file. Returns list of dicts."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Names"]

    mailboxes = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        email = ws.cell(r, 1).value
        pwd = ws.cell(r, 2).value or "Exz@2631"
        fname = ws.cell(r, 3).value or ""
        lname = ws.cell(r, 4).value or ""
        company = ws.cell(r, 5).value or ""
        website = ws.cell(r, 6).value or ""
        phone = ws.cell(r, 7).value or ""
        company_email = ws.cell(r, 8).value or ""

        if not email:
            # Fallback: construct from first name
            if fname:
                email = f"{fname.strip().lower()}@exzelon.in"
            else:
                continue

        email = email.strip().lower()
        if email in seen:
            continue
        seen.add(email)

        mailboxes.append({
            "email": email,
            "password": pwd.strip(),
            "first_name": fname.strip(),
            "last_name": lname.strip(),
            "company": company.strip(),
            "website": website.strip(),
            "phone": phone.strip() if isinstance(phone, str) else str(phone),
            "company_email": company_email.strip(),
        })

    return mailboxes


def create_mailbox(headers: dict, mb: dict) -> dict:
    """Create a single mailbox via API. Returns response dict."""
    mail_server = get_mail_server(mb["email"])
    display_name = f"{mb['first_name']} {mb['last_name']}".strip()

    payload = {
        "email": mb["email"],
        "password": mb["password"],
        "display_name": display_name,
        "sender_first_name": mb["first_name"],
        "sender_last_name": mb["last_name"],
        "provider": DEFAULT_PROVIDER,
        "auth_method": DEFAULT_AUTH_METHOD,
        "smtp_host": mail_server,
        "smtp_port": DEFAULT_SMTP_PORT,
        "imap_host": mail_server,
        "imap_port": DEFAULT_IMAP_PORT,
        "daily_send_limit": DAILY_SEND_LIMIT,
        "is_active": True,
        "warmup_status": "inactive",
    }

    resp = requests.post(f"{API_BASE}/mailboxes", headers=headers,
                         json=payload, timeout=30)
    return {"status": resp.status_code, "data": resp.json()}


def test_connection(headers: dict, mailbox_id: int) -> dict:
    """Test connection for an existing mailbox."""
    resp = requests.post(f"{API_BASE}/mailboxes/{mailbox_id}/test-connection",
                         headers=headers, timeout=60)
    return resp.json()


def main():
    print("=" * 60)
    print("BULK MAILBOX IMPORT")
    print("=" * 60)

    # 1. Login
    print("\n[1/4] Authenticating...")
    token = login()
    headers = {"Authorization": f"Bearer {token}"}
    print("  Authenticated OK")

    # 2. Get existing mailboxes
    print("\n[2/4] Checking existing mailboxes...")
    existing = get_existing_mailboxes(headers)
    print(f"  Found {len(existing)} existing mailboxes")

    # 3. Read Excel
    print("\n[3/4] Reading Excel file...")
    mailboxes = read_excel()
    print(f"  Found {len(mailboxes)} unique mailboxes in Excel")

    # Domain breakdown
    domains = Counter(m["email"].split("@")[1] for m in mailboxes)
    for domain, count in sorted(domains.items()):
        print(f"    {domain}: {count}")

    # Filter out existing
    to_add = [m for m in mailboxes if m["email"] not in existing]
    skipped = len(mailboxes) - len(to_add)
    if skipped:
        print(f"  Skipping {skipped} already existing mailboxes")
    print(f"  Will add {len(to_add)} new mailboxes")

    if not to_add:
        print("\nNothing to add!")
        return

    # 4. Create mailboxes
    print(f"\n[4/4] Creating {len(to_add)} mailboxes...")
    created = 0
    failed = 0
    created_ids = []
    errors = []

    for i, mb in enumerate(to_add):
        try:
            result = create_mailbox(headers, mb)
            if result["status"] in (200, 201):
                created += 1
                mid = result["data"].get("mailbox_id", result["data"].get("id"))
                if mid:
                    created_ids.append(mid)
                if created % 50 == 0:
                    print(f"  Created {created}/{len(to_add)}...")
            else:
                failed += 1
                detail = result["data"].get("detail", str(result["data"]))
                if isinstance(detail, list):
                    detail = "; ".join(str(d) for d in detail)
                errors.append(f"{mb['email']}: {detail[:100]}")
                if failed <= 5:
                    print(f"  FAIL: {mb['email']} -> {detail[:80]}")

            # Re-auth every 200 requests (tokens expire in 30min)
            if (i + 1) % 200 == 0:
                try:
                    token = login()
                    headers = {"Authorization": f"Bearer {token}"}
                    print(f"  Re-authenticated at {i+1}/{len(to_add)}")
                except Exception:
                    pass

        except requests.exceptions.RequestException as e:
            failed += 1
            errors.append(f"{mb['email']}: {type(e).__name__}")
            if failed <= 5:
                print(f"  ERROR: {mb['email']} -> {e}")
            time.sleep(1)

    print(f"\n{'='*60}")
    print(f"RESULTS")
    print(f"{'='*60}")
    print(f"  Total in Excel: {len(mailboxes)}")
    print(f"  Already existed: {skipped}")
    print(f"  Created: {created}")
    print(f"  Failed: {failed}")
    if errors:
        print(f"\n  First {min(len(errors), 20)} errors:")
        for e in errors[:20]:
            print(f"    - {e}")

    # Write results
    with open("bulk_mailbox_results.json", "w") as f:
        json.dump({
            "total": len(mailboxes),
            "skipped": skipped,
            "created": created,
            "failed": failed,
            "created_ids": created_ids,
            "errors": errors[:50],
        }, f, indent=2)
    print(f"\nResults saved to bulk_mailbox_results.json")


if __name__ == "__main__":
    main()
